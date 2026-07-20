"""Excel export utilities for customer records."""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


EXCEL_FILE_NAME = "customer_report.xlsx"
WORKSHEET_NAME = "Customer Report"

COLUMN_MAPPING = {
    "contact_person_name": "Customer Name",
    "email_id": "Email ID",
    "organisation_name": "Organisation",
    "mobile_number": "Mobile Number",
    "input_source": "Source",
    "designation": "Designation",
    "address": "Address",
    "subject": "Subject",
    "extraction_confidence": "Extraction Confidence",
    "duplicate_status": "Duplicate Status",
    "confidence_score": "Confidence Score",
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
EVEN_ROW_FILL = PatternFill(fill_type="solid", fgColor="F7FBFF")
ODD_ROW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
HEADER_FONT = Font(bold=True)


def _build_export_rows(customers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Convert customer dictionaries into Excel-ready rows."""
    rows: list[dict[str, Any]] = []

    for customer in customers:
        rows.append(
            {
                column_name: customer.get(source_key, "")
                for source_key, column_name in COLUMN_MAPPING.items()
            }
        )

    return rows


def _auto_adjust_column_widths(worksheet: Any) -> None:
    """Set worksheet column widths based on cell content."""
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))

        worksheet.column_dimensions[column_letter].width = min(max_length + 3, 45)


def _format_worksheet(worksheet: Any) -> None:
    """Apply header, freeze pane, and row formatting to the worksheet."""
    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        fill = EVEN_ROW_FILL if row_index % 2 == 0 else ODD_ROW_FILL
        for cell in row:
            cell.fill = fill

    _auto_adjust_column_widths(worksheet)


def export_customers_to_excel(customers: list[dict[str, Any]]) -> BytesIO:
    """Return a formatted Excel report for customer records as a BytesIO object."""
    if not isinstance(customers, list):
        raise TypeError("customers must be a list of dictionaries")

    for index, customer in enumerate(customers):
        if not isinstance(customer, dict):
            raise TypeError(f"customer at index {index} must be a dictionary")

    output = BytesIO()
    dataframe = pd.DataFrame(_build_export_rows(customers), columns=COLUMN_MAPPING.values())

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=WORKSHEET_NAME)
        worksheet = writer.sheets[WORKSHEET_NAME]
        _format_worksheet(worksheet)

    output.seek(0)
    return output

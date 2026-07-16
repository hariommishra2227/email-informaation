"""Verification script for the Excel exporter."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from excel_exporter import (
    EXCEL_FILE_NAME,
    WORKSHEET_NAME,
    export_customers_to_excel,
)


def build_sample_customers() -> list[dict[str, object]]:
    """Create at least 20 sample customer records for export verification."""
    statuses = ["Unique", "Duplicate", "Possible Duplicate", "Unique"]
    customers: list[dict[str, object]] = []

    for index in range(1, 21):
        status = statuses[(index - 1) % len(statuses)]
        customers.append(
            {
                "contact_person_name": f"Customer {index}",
                "email_id": f"customer{index}@example.com",
                "organisation_name": f"Organisation {((index - 1) % 5) + 1}",
                "mobile_number": f"+91 98765432{index:02d}",
                "input_source": "PDF" if index % 2 else "Manual Paste",
                "designation": "Manager" if index % 2 else "Director",
                "address": f"{index} Market Street, City",
                "subject": f"Quotation Request {index}",
                "extraction_confidence": 100,
                "duplicate_status": status,
                "confidence_score": 100 if status == "Duplicate" else 88 if status == "Possible Duplicate" else 0,
            }
        )

    return customers


def test_export_customers_to_excel() -> None:
    """Export sample customers and validate workbook structure and formatting."""
    excel_buffer = export_customers_to_excel(build_sample_customers())
    output_path = Path(EXCEL_FILE_NAME)
    output_path.write_bytes(excel_buffer.getvalue())

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [WORKSHEET_NAME]

    worksheet = workbook[WORKSHEET_NAME]
    assert worksheet.max_row == 21
    assert worksheet.max_column == 11
    assert worksheet.freeze_panes == "A2"
    assert worksheet["A1"].value == "Customer Name"
    assert worksheet["D2"].value == "+91 9876543201"
    assert worksheet["E1"].value == "Source"
    assert worksheet["A1"].font.bold is True
    assert worksheet["A1"].fill.fgColor.rgb == "00D9EAF7"
    assert worksheet["A2"].fill.fgColor.rgb == "00F7FBFF"
    assert worksheet["A3"].fill.fgColor.rgb == "00FFFFFF"
    assert worksheet.column_dimensions["A"].width > len("Customer Name")


def main() -> None:
    """Create and validate a sample customer Excel report."""
    test_export_customers_to_excel()
    print(f"Created {EXCEL_FILE_NAME} with 20 customer records.")


if __name__ == "__main__":
    main()

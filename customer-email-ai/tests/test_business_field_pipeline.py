"""Regression coverage for the single eight-field business pipeline."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook

import config
from duplicate_handler import customer_record_to_contact
from models import CustomerRecord, OutlookMessage
from repository import EmailSyncRepository
from services.email_processor import build_customer_record
from services.export_service import cleanup_export, create_large_csv_export, create_large_excel_export
from storage import database
from sync import OutlookMailboxSyncService


HEADERS = ["Client Name", "Contact Person Name", "Contact Email", "Phone Number", "Full Address", "Location", "Subject", "Email Date"]


@pytest.fixture()
def isolated_db(monkeypatch: pytest.MonkeyPatch) -> None:
    database.DATABASE_PATH = Path(":memory:")
    monkeypatch.setattr(config, "DATABASE_URL", "")
    monkeypatch.setattr(config, "APP_ENV", "test")
    monkeypatch.setattr(config, "LLM_ENABLED", False)
    database.initialize_database()


class CompleteEngine:
    def extract(self, text: str, **kwargs):
        return {
            "contact_person_name": "Extracted Person", "organisation_name": "Example Ltd",
            "email_id": "body@example.net", "mobile_number": "+91 98765 43210",
            "address": "12 Business Park, Pune, Maharashtra 411001", "location": "Pune",
            "name_confidence": .9, "email_confidence": .9, "organisation_confidence": .9,
            "address_confidence": .9,
        }


def _message(sender_email: str = "person@example.com") -> OutlookMessage:
    return OutlookMessage(
        message_id="m1", user_id="u", sender_name="", sender_email=sender_email,
        subject="Graph Subject", body="Location: Pune", received_datetime="2026-08-07T10:30:00Z",
        is_read=False,
    )


def test_manual_record_preserves_all_business_fields(isolated_db: None) -> None:
    message = _message()
    record = build_customer_record(
        user_id="u", text=message.body, source="Outlook", source_message_id=message.message_id,
        sender_email=message.sender_email, sender_name=message.sender_name, subject=message.subject,
        received_datetime=message.received_datetime, engine=CompleteEngine(),
    )
    assert (record.organisation, record.contact_name, record.email, record.mobile) == (
        "Example Ltd", "Extracted Person", "person@example.com", "+91 98765 43210"
    )
    assert (record.address, record.location, record.subject, record.email_date) == (
        "12 Business Park, Pune, Maharashtra 411001", "Pune", "Graph Subject", "07-Aug-2026 10:30"
    )


def test_background_sync_preserves_graph_fields_and_filters_internal(isolated_db: None, monkeypatch: pytest.MonkeyPatch) -> None:
    messages = [_message(), OutlookMessage(**{**_message("staff@itsipl.com").to_dict(), "message_id": "internal"})]
    monkeypatch.setattr("services.graph_client.iter_mailbox_message_pages", lambda *a, **k: [messages])
    result = OutlookMailboxSyncService(EmailSyncRepository()).sync_mailbox("u", engine=CompleteEngine())
    row = database.list_customers("u")[0]
    assert result.processed_emails == 1 and result.skipped_emails == 1
    assert [row[key] for key in ("organisation", "contact_name", "email", "mobile", "address", "location", "subject", "email_date")] == [
        "Example Ltd", "Extracted Person", "person@example.com", "+91 98765 43210",
        "12 Business Park, Pune, Maharashtra 411001", "Pune", "Graph Subject", "07-Aug-2026 10:30",
    ]


def test_conversion_repository_and_merge_preserve_business_fields(isolated_db: None) -> None:
    database.insert_customer(CustomerRecord(user_id="u", email="a@example.com", normalized_email="a@example.com", organisation="Good Org"))
    incoming = CustomerRecord(
        user_id="u", contact_name="A", email="a@example.com", normalized_email="a@example.com",
        location="Pune", subject="Subject", email_date="07-Aug-2026 10:30", source_message_id="m",
    )
    EmailSyncRepository().upsert_contact(customer_record_to_contact(incoming), user_id="u")
    row = database.list_customers("u")[0]
    assert row["organisation"] == "Good Org"
    assert (row["location"], row["subject"], row["email_date"]) == ("Pune", "Subject", "07-Aug-2026 10:30")


def test_excel_and_csv_have_exact_business_schema(isolated_db: None) -> None:
    database.insert_customer(CustomerRecord(user_id="u", email="a@example.com", location="Pune", subject="S", email_date="D"))
    excel_path = create_large_excel_export("u")
    csv_path = create_large_csv_export("u")
    try:
        assert list(load_workbook(excel_path, read_only=True).active.values)[0] == tuple(HEADERS)
        with csv_path.open(encoding="utf-8-sig", newline="") as handle:
            assert next(csv.reader(handle)) == HEADERS
    finally:
        cleanup_export(excel_path)
        cleanup_export(csv_path)


def test_customer_excel_exports_multiple_processed_outlook_records_only(isolated_db: None) -> None:
    database.insert_customer(CustomerRecord(user_id="u", email="one@example.com", source="Outlook"))
    database.insert_customer(CustomerRecord(user_id="u", email="two@example.com", source="Outlook"))
    database.insert_customer(CustomerRecord(user_id="u", email="manual@example.com", source="Manual"))

    excel_path = create_large_excel_export("u", source="Outlook")
    try:
        rows = list(load_workbook(excel_path, read_only=True).active.values)
    finally:
        cleanup_export(excel_path)

    assert rows[0] == tuple(HEADERS)
    assert [row[2] for row in rows[1:]] == ["one@example.com", "two@example.com"]


def test_file_database_does_not_reuse_prior_in_memory_state(
    isolated_db: None, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    database.insert_customer(CustomerRecord(user_id="u", email="old@example.com"))
    monkeypatch.setattr(database, "DATABASE_PATH", tmp_path / "isolated.db")
    database.initialize_database()
    assert database.list_customers("u") == []


def test_loaded_message_status_lookup_is_not_limited_to_first_100(isolated_db: None) -> None:
    messages = [
        OutlookMessage(
            message_id=f"m-{index}", user_id="u", sender_name="", sender_email=f"p{index}@example.com",
            subject="", body="", received_datetime=f"2026-08-07T10:{index % 60:02d}:00Z", is_read=False,
        )
        for index in range(125)
    ]
    for message in messages:
        database.upsert_outlook_message(message)
        database.set_message_status("u", message.message_id, "Unique")
    statuses = database.message_statuses_for_ids("u", [message.message_id for message in messages])
    assert len(statuses) == 125
    assert statuses["m-124"] == "Unique"

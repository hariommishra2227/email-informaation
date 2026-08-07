"""Regression tests for AWS production-readiness blockers."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import insert

import config
from database.models import ExtractedContact
from models import CustomerRecord, OutlookMessage
from services.export_service import cleanup_export, create_large_excel_export
from services.token_cache_crypto import PREFIX, decrypt_cache, encrypt_cache
from storage import database
from workers.job_handler import enqueue_email_sync


@pytest.fixture()
def isolated_db(monkeypatch: pytest.MonkeyPatch) -> None:
    database.DATABASE_PATH = Path(":memory:")
    monkeypatch.setattr(config, "DATABASE_URL", "")
    monkeypatch.setattr(config, "APP_ENV", "test")
    database.initialize_database()


def test_review_functions_use_orm_tables(isolated_db: None) -> None:
    record_id = database.insert_customer(CustomerRecord(user_id="u", contact_name="Old", email="a@example.com"))
    database.update_customer_review(record_id, {"contact_name": "New", "review_status": "Approved"}, reviewed_by="r", notes="ok")

    row = database.list_customers("u")[0]
    audit = database.list_review_audit(record_id)

    assert row["contact_name"] == "New"
    assert row["name_source"] == "manual_review"
    assert row["review_status"] == "Approved"
    assert audit[0]["old_value"] == "Old"
    assert audit[0]["new_value"] == "New"


def test_large_export_iterates_beyond_ui_page_cap(isolated_db: None, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(config, "EXPORT_CHUNK_SIZE", 250)
    for index in range(1001):
        database.insert_customer(
            CustomerRecord(
                user_id="u",
                contact_name=f"Person {index}",
                email=f"person{index}@example.com",
                organisation=f"Org {index}",
            )
        )

    path = create_large_excel_export("u", limit=1001)
    try:
        rows = list(load_workbook(path, read_only=True).active.iter_rows(values_only=True))
    finally:
        cleanup_export(path)

    assert list(rows[0]) == [
        "Client Name", "Contact Person Name", "Contact Email", "Phone Number",
        "Full Address", "Location", "Subject", "Email Date",
    ]
    email_column = rows[0].index("Contact Email")
    exported_emails = [row[email_column] for row in rows[1:]]
    assert len(exported_emails) == 1001
    assert len(set(exported_emails)) == 1001
    assert "person1000@example.com" in exported_emails


def test_email_contact_relationship_is_idempotent(isolated_db: None) -> None:
    message = OutlookMessage(
        message_id="m1",
        user_id="u",
        sender_name="A",
        sender_email="a@example.com",
        subject="Hi",
        body="Email: a@example.com",
        received_datetime="2026-08-03T00:00:00Z",
        is_read=False,
    )
    database.upsert_outlook_message(message)
    first = database.insert_customer(CustomerRecord(user_id="u", email="a@example.com", normalized_email="a@example.com", source_message_id="m1"))
    second = database.insert_customer(CustomerRecord(user_id="u", email="a@example.com", normalized_email="a@example.com", source_message_id="m1"))

    assert first == second


def test_token_cache_encryption_round_trip_and_plaintext_migration(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(config, "TOKEN_CACHE_ENCRYPTION_KEY", "dev-test-key")
    encrypted = encrypt_cache('{"token": "secret"}')

    assert encrypted.startswith(PREFIX)
    assert decrypt_cache(encrypted) == ('{"token": "secret"}', False)
    assert decrypt_cache('{"legacy": true}') == ('{"legacy": true}', True)


def test_enqueue_prevents_two_active_sync_jobs(isolated_db: None, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(config, "JOB_BACKEND", "sqs")
    monkeypatch.setattr(config, "AWS_SQS_QUEUE_URL", "https://sqs.example/queue")

    class FakeSqs:
        def send_message(self, **kwargs):
            return {"MessageId": "1"}

    class FakeBoto:
        def client(self, *args, **kwargs):
            return FakeSqs()

    monkeypatch.setitem(__import__("sys").modules, "boto3", FakeBoto())

    first = enqueue_email_sync("u")
    second = enqueue_email_sync("u")

    assert first["status"] == "queued"
    assert second["existing"] is True


@pytest.mark.performance
def test_100000_record_export_performance_smoke(isolated_db: None, monkeypatch: pytest.MonkeyPatch) -> None:
    if not bool(int(__import__("os").environ.get("RUN_100K_EXPORT_TEST", "0"))):
        pytest.skip("Set RUN_100K_EXPORT_TEST=1 to execute the 100,000-record export smoke test.")
    monkeypatch.setattr(config, "EXPORT_CHUNK_SIZE", 5000)
    with database.db_session() as session:
        mailbox = database.ensure_mailbox(session, "u")
        for start in range(0, 100000, 5000):
            session.execute(
                insert(ExtractedContact),
                [
                    {
                        "mailbox_id": mailbox.id,
                        "email_address": f"p{index}@example.com",
                        "normalized_email": f"p{index}@example.com",
                        "source": "Outlook",
                    }
                    for index in range(start, start + 5000)
                ],
            )
    path = create_large_excel_export("u", limit=100000)
    try:
        assert path.exists()
        assert path.stat().st_size > 0
    finally:
        cleanup_export(path)
